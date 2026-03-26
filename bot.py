from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import load_workbook, Workbook

kullanici = {}

def ilana_yaz(urun, tonaj, fiyat, konum):
    try:
        wb = load_workbook("ilanlar.xlsx")
        sheet = wb.active
    except:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Ürün", "Tonaj", "Fiyat", "Konum"])

    sheet.append([urun, tonaj, fiyat, konum])
    wb.save("ilanlar.xlsx")

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("İlan vermek için 'ilan' yaz")

async def mesaj(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user_id = update.message.from_user.id

    if text.lower() == "ilan":
        kullanici[user_id] = {"step": "urun"}
        await update.message.reply_text("Ürün?")
        return

    if user_id in kullanici:
        step = kullanici[user_id]["step"]

        if step == "urun":
            kullanici[user_id]["urun"] = text
            kullanici[user_id]["step"] = "tonaj"
            await update.message.reply_text("Tonaj?")
            return

        elif step == "tonaj":
            kullanici[user_id]["tonaj"] = text
            kullanici[user_id]["step"] = "fiyat"
            await update.message.reply_text("Fiyat?")
            return

        elif step == "fiyat":
            kullanici[user_id]["fiyat"] = text
            kullanici[user_id]["step"] = "konum"
            await update.message.reply_text("Konum?")
            return

        elif step == "konum":
            kullanici[user_id]["konum"] = text

            ilana_yaz(
                kullanici[user_id]["urun"],
                kullanici[user_id]["tonaj"],
                kullanici[user_id]["fiyat"],
                text
            )

            await update.message.reply_text("✅ Kaydedildi")

            del kullanici[user_id]
            return

app = ApplicationBuilder().token("8296117624:AAEc7-nbhRl0Fy_-XgBaEDQFvcbSv43tSJU").build()

app.add_handler(CommandHandler("start", start))
app.add_handler(MessageHandler(filters.TEXT, mesaj))

print("ÇALIŞIYOR")
app.run_polling()